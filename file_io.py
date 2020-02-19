# -*- coding: utf-8 -*-
# @Author: Jie Yang
# @Date:   2019-01-24 13:03:23
# @Last Modified by:   Jie Yang,     Contact: jieynlp@gmail.com
# @Last Modified time: 2020-02-19 11:25:48
# -*- coding: utf-8 -*-
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import numpy as np
import csv
from utils import clinic_text_processing, filter_duplicacte
import random
import copy

seed_num = 42
random.seed(seed_num)
np.random.seed(seed_num)

def load_text_classification_data_txt(input_file, split_token=" ||| ", shuffle=False, filter_duplicacte_instances=True):
    r''' load text classification data in normal text format (not limited in .txt file)
        Args:
            input_file (string): input text file directory
            split_token (string): split the text based on split_token
        Return:
            x (list): the text list of all documents/sentences
            y (list): the label list of the corresponding documents/sentences. = [] when no split_token found in text.
    '''
    print("Start loading txt file from %s."%input_file)
    x = []
    y = []
    with open(input_file,'rb') as infile:
        fins = infile.readlines()
        include_label = True 
        if split_token not in fins[0]:
            include_label = False
        for line in fins:
            if include_label:
                pair = line.strip('\r\n').split(split_token)
                assert(len(pair)==2)
                x.append(pair[0])
                y.append(int(pair[1]))
            else:
                x.append(line.strip('\r\n'))
    if filter_duplicacte_instances:
        x, y = filter_duplicacte([x,y])
    instance_num = len(x)
    print("\tInstance Num: %s"%instance_num)
    if shuffle:
        combined = list(zip(x, y))
        random.shuffle(combined)
        x[:], y[:] = zip(*combined)
        print("\tInstances shuffled.")
    return x, y


def load_text_classification_data_csv(input_report, column_list, shuffle=False, filter_duplicacte_instances=True, include_title=True):
    r''' load the xlsx train data: 9K_reports_20181228.xlsx
        Args:
            input_report (string): input xlsx file directory
            X_id (string): the column id of input X ('A', 'B',...)
            Y_id (string): the column id of label Y ('A', 'B',...)
        Return:
            Descriptions, ADE, ADR, HSR, MRN: all lists, corresponding information, text is preprocessed.
    '''
    print("Start loading csv file from %s."%input_report)
    output_list = [[] for a in column_list]
    column_num = len(column_list)
    with open(input_report,'r') as csv_file:
        csv_reader = csv.reader(csv_file)
        if include_title:
            next(csv_reader, None)
        for row in csv_reader:
            for idx in range(column_num):
                the_x = row[column_list[idx]]
                # the_x = unicode(the_x, errors='ignore')
                the_x = clinic_text_processing(the_x)
                output_list[idx].append(the_x)
    if filter_duplicacte_instances:
        output_list = filter_duplicacte(output_list, '\t', -1)
    return output_list




def load_text_classification_data_xlsx(input_report, X_id, Y_id, shuffle=False, filter_duplicacte_instances=True, include_title=True):
    r''' load the xlsx train data: 9K_reports_20181228.xlsx
        Args:
            input_report (string): input xlsx file directory
            X_id (string): the column id of input X ('A', 'B',...)
            Y_id (string): the column id of label Y ('A', 'B',...)
        Return:
            Descriptions, ADE, ADR, HSR, MRN: all lists, corresponding information, text is preprocessed.
    '''
    print("Start loading xlsx file from %s."%input_report)
    report = load_workbook(input_report)
    active_workbook = report.active
    print("\tworkbook: %s"%(report.sheetnames))
    ws = report[report.sheetnames[0]]
    instance_num = len(ws['A']) - 1
    X = []
    Y = []
    for idx in range(1, instance_num+1):
        the_x = ws[X_id+str(idx)].value
        the_y = str(ws[Y_id+str(idx)].value)
        if the_y in ['0','1']:
            the_x = clinic_text_processing(the_x)
            X.append(the_x)
            Y.append(int(the_y))

    if filter_duplicacte_instances:
        X,Y = filter_duplicacte([X,Y])
    instance_num = len(X)
    print("\tInstance Num: %s"%instance_num)
    if shuffle:
        combined = list(zip(X,Y))
        random.shuffle(combined)
        X[:],Y[:] = zip(*combined)
        print("\tInstances shuffled.")
    print("File loaded.")
    return X, Y


def load_multi_text_classification_data_xlsx(input_report, X_id, Y_id, new_Y_id, shuffle=False, filter_duplicacte_instances=True, include_title=True):
    r''' load the xlsx train data: 9K_reports_20181228.xlsx
        Args:
            input_report (string): input xlsx file directory
            X_id (string): the column id of input X ('A', 'B',...)
            Y_id (string): the column id of label Y ('A', 'B',...)
        Return:
            Descriptions, ADE, ADR, HSR, MRN: all lists, corresponding information, text is preprocessed.
    '''
    print("Start loading xlsx file from %s."%input_report)
    report = load_workbook(input_report)
    active_workbook = report.active
    print("\tworkbook: %s"%(report.sheetnames))
    ws = report[report.sheetnames[0]]
    instance_num = len(ws['A']) - 1
    X = []
    Y = []
    new_Y = []
    for idx in range(1, instance_num+1):
        the_x = ws[X_id+str(idx)].value
        the_y = str(ws[Y_id+str(idx)].value)
        the_new_y = str(ws[new_Y_id+str(idx)].value)
        if the_y in ['0','1'] and the_new_y in ['0','1']:
            the_x = clinic_text_processing(the_x)
            X.append(the_x)
            Y.append(int(the_y))
            # print(the_new_y)
            new_Y.append(int(the_new_y))
    if filter_duplicacte_instances:
        X,Y, new_Y = filter_duplicacte([X,Y, new_Y])
    instance_num = len(X)
    print("\tInstance Num: %s"%instance_num)
    if shuffle:
        combined = list(zip(X,Y,new_Y))
        random.shuffle(combined)
        X[:],Y[:], new_Y[:] = zip(*combined)
        print("\tInstances shuffled.")
    print("File loaded.")
    return X, Y, new_Y


def xlsx_extract_column(input_report, id_list, include_title=True):
    r''' load the xlsx train data: 9K_reports_20181228.xlsx
        Args:
            input_report (string): input xlsx file directory
            X_id (string): the column id of input X ('A', 'B',...)
            Y_id (string): the column id of label Y ('A', 'B',...)
        Return:
            Descriptions, ADE, ADR, HSR, MRN: all lists, corresponding information, text is preprocessed.
    '''
    print("Start loading xlsx file from %s."%input_report)
    report = load_workbook(input_report)
    active_workbook = report.active
    print("\tworkbook: %s"%(report.sheetnames))
    ws = report[report.sheetnames[0]]
    instance_num = len(ws['A']) - 1
    column_num = len(id_list)
    output_list = [[] for idx in range(column_num)]
    new_Y = []
    start_row = 1
    print("Extract column num:", column_num)
    if include_title:
        start_row = 2
    for idx in range(start_row, instance_num+2):
        for idy in range(column_num):
            output_list[idy].append(ws[id_list[idy]+str(idx)].value)
    print("Total row number:", len(output_list[0]))
    print("File loaded.")
    return output_list


def load_pair_txt_file(input_file):
    r''' load the new data to be decoded: report_from_final-annotated.txt
        Args:
            input_file (string): input file directory
        Return:
            record_id_list (list)：id list of document/sentence
            record_list (list):  content list of document/sentence, with preprocessing
    '''
    
    
    fins = open(input_file,'r', errors='ignore').readlines()
    print("Original line num:", len(fins))
    record_id_list = []
    record_list = []
    for line in fins:
        line = line.strip('\n')
        if '\t' not in line:
            continue
        pair = line.split('\t',1)
        content = clinic_text_processing(pair[1])
        if len(content) < 1:
            continue
        record_id_list.append(pair[0])
        record_list.append(content)
    filter_duplicacte_instances = True
    if filter_duplicacte_instances:
        record_id_list,record_list = filter_duplicacte([record_id_list,record_list])
    print("Pair txt file loaded, source: %s"%input_file)
    return record_id_list, record_list


def write_decode_to_csv(csv_file, X, Y, title_list=[]):
    r''' write the decoded label with input to csv file
        Args:
            csv_file (string): output file directory
            X (list): list of input X
            Y (list): list of labels
    '''
    instance_num  = len(X)
    assert(instance_num==len(Y))
    with open(csv_file, 'w') as cfile:
        csvwriter = csv.writer(cfile) 
        if title_list:
            csvwriter.writerow(title_list)
        else:
            csvwriter.writerow(['Label',"Input"])
        for the_x, the_y in zip(X,Y):
            csvwriter.writerow([the_y, the_x])
    print("Decoded results has been saved into file %s"%csv_file)


def write_multi_decode_to_csv(csv_file, X, Y_list, title_list=[]):
    r''' write the decoded label with input to csv file
        Args:
            csv_file (string): output file directory
            X (list): list of input X
            Y_list (list): list of labels for different models
            title_list (list of string): name of differnet models
    '''
    model_num = len(Y_list)
    instance_num  = len(X)
    print(instance_num,len(Y_list[0]) )
    assert(instance_num==len(Y_list[0]))
    with open(csv_file, 'w') as cfile:
        csvwriter = csv.writer(cfile) 
        csvwriter.writerow(title_list+['Input'])
        for idx in range(instance_num):
            decode_list = []
            for idy in range(model_num):
                decode_list.append(Y_list[idy][idx])
            decode_list.append(X[idx])
            csvwriter.writerow(decode_list)
    print("Decoded results has been saved into file %s"%csv_file)


def select_medium_to_annotate(input_file, output_file):
    ## load decoded data from four different models.
    
    output_list = load_text_classification_data_csv(input_file, [0,1,2,3,4], False, False, include_title=True)
    new_description_dict = {}
    for idx in range(len(output_list)-1):
        new_des = copy.deepcopy(output_list[-1])
        score, des = (list(t) for t in zip(*sorted(zip(output_list[idx], new_des), reverse=True)))
        medium_id = -1 
        for idy in range(len(score)):
            if medium_id < 0:
                if float(score[idy]) < 0.5:
                    medium_id = idy
        print("Medium id:", medium_id)
        count = 0
        for idy in range(medium_id-50, medium_id+50):
            if 'rash' in des[idy]:
                count +=1
            if des[idy] not in new_description_dict:
                new_description_dict[des[idy]] = 1 
        print(count)

    print(len(new_description_dict))
    # exit(0)
    new_list = new_description_dict.keys()
    fout = open(output_file, 'wb')
    csvwriter = csv.writer(fout)
    csvwriter.writerow(["ADE", "ADR", "HSR","Description"])
    for each in new_list:
        csvwriter.writerow(["","","",each])

def select_top_to_annotate(input_file, output_file):
    ## load decoded data from four different models.
    
    output_list = load_text_classification_data_csv(input_file, [1,2,3,4,6], False, False, include_title=True)
    new_description_dict = {}
    for idx in range(len(output_list)-1):
        new_des = copy.deepcopy(output_list[-1])
        score, des = (list(t) for t in zip(*sorted(zip(output_list[idx], new_des), reverse=True)))
        count = 0
        for idy in range(100):
            if 'rash' in des[idy]:
                count +=1
            if des[idy] not in new_description_dict:
                new_description_dict[des[idy]] = 1 
        print(count)

    print(len(new_description_dict))
    # exit(0)
    new_list = new_description_dict.keys()
    fout = open(output_file, 'wb')
    csvwriter = csv.writer(fout)
    csvwriter.writerow(["ADE", "ADR", "HSR","Description"])
    for each in new_list:
        csvwriter.writerow(["","","",each])

def select_low_to_annotate(input_file, output_file):
    ## load decoded data from four different models. only work for sonm project
    
    output_list = load_text_classification_data_csv(input_file, [5,6,7,8,9], False, True, include_title=True)
    new_out_list = []
    # for a,b,c,d,e in zip(output_list[0], output_list[1],output_list[2],output_list[3],output_list[4]):


    new_description_dict = {}
    for idx in range(len(output_list)-1):
        new_des = copy.deepcopy(output_list[-1])
        score, des = (list(t) for t in zip(*sorted(zip(output_list[idx], new_des), reverse=False)))
        for idy in range(100):
            if des[idy] not in new_description_dict:
                new_description_dict[des[idy]] = 1 
    print(len(new_description_dict))
    # exit(0)
    new_list = new_description_dict.keys()
    fout = open(output_file, 'wb')
    csvwriter = csv.writer(fout)
    csvwriter.writerow(["Label", "Text"])
    for each in new_list:
        csvwriter.writerow(["",each])

def select_filter_train_to_annotate(input_file, train_file, output_file):
    ## load decoded data from four different models. only work for sonm project
    Y_old, X = load_text_classification_data_csv(train_file, [0,1], True)
    train_dict = {}
    for x in X:
        if x not in train_dict:
            train_dict[x] = 1

    output_list = load_text_classification_data_csv(input_file, [5,6,7,8,9], False, True, include_title=True)
    first_line = ["svm","logistic","randomforest","xgboost","Text"]
    origin_num = len(output_list[0])
    fout = open(output_file, 'wb')
    csvwriter = csv.writer(fout)
    csvwriter.writerow(first_line)
    out_num = len(output_list)
    print("Old:", origin_num)
    new_count = 0
    for idx in range(origin_num):
        if output_list[-1][idx] not in train_dict:
            csvwriter.writerow([output_list[idy][idx] for idy in range(out_num-1)]+[output_list[-1][idx]])
            new_count += 1
    print("New:", new_count)
   

def count_avg_sentence_length(input_xlsx, column):
    text_list = xlsx_extract_column(input_xlsx, [column])[0]
    total_length = 0 
    for each_text in text_list:
        words = each_text.strip('\n').split()
        total_length += len(words)
    avg_length = (total_length+0.)/len(text_list)

    print("Sent num:", len(text_list), "Avg length: ", avg_length)


if __name__ == '__main__':
    print("file_io.py")
    
